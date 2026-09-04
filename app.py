import os, io, json, base64, pickle, sqlite3, uuid, hashlib
from datetime import datetime
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError
import pandas as pd
import streamlit as st
import firebase_admin
from firebase_admin import credentials, firestore, auth as firebase_auth

st.set_page_config(page_title='Gestão Almoxarifado | Inventário Rotativo', page_icon='📦', layout='wide', initial_sidebar_state='expanded')
DATA=os.path.join(os.path.dirname(__file__),'inventario_operacional.sqlite3')

DEFAULT={
 'theme':'Dark','font':'Arial','font_size':16,'title_size':31,
 'primary':'#FFD63B','hover':'#F7C928','icon_color':'#FFD63B','dark_bg':'#080B0A','dark_panel':'#101614','dark_panel2':'#141A17','dark_border':'#2B3732','dark_text':'#F4F5F2','dark_muted':'#A9B1AC',
 'clean_bg':'#F5F6F4','clean_panel':'#FFFFFF','clean_panel2':'#F0F2EF','clean_border':'#D8DDD9','clean_text':'#161A18','clean_muted':'#626B66',
 'title':'GESTÃO ALMOXARIFADO','subtitle':'01 · ACURÁCIA DE ESTOQUE  |  Inventário Rotativo','sidebar_sub':'SISTEMA OPERACIONAL DE ESTOQUE','menu':'MENU',
 'dash':'DASHBOARD','inv':'INVENTÁRIO ROTATIVO','db':'BANCO DE DADOS','reg':'REGISTRO','report':'REPORTAR INCONSISTÊNCIAS','settings':'CONFIGURAÇÕES',
 'sidebar_width':250,'menu_gap':2,'report_top':0,'logo_w':190,'logo_h':70,'logo_align':'center','logo_top':-10,'sub_top':0,'menu_top':0,'sidebar_align':'left','sidebar_font':12,'item_h':42,'gap':8,'dash_top':0,'inv_top':0,'db_top':0,'reg_top':0,'settings_top':0,'show_footer':True,
 'blind_default':False,'dashboard_title':'Dashboard','inventory_title':'Inventário Rotativo','database_title':'Banco de Dados','register_title':'Registro','dashboard_subtitle':'Visão geral dos indicadores do estoque.','inventory_subtitle':'Controle e execução dos inventários rotativos.','database_subtitle':'Importação, tratamento e classificação da base de estoque.','register_subtitle':'Histórico dos inventários e das contagens realizadas.'
}

def dbconn():
 c=sqlite3.connect(DATA); c.execute('CREATE TABLE IF NOT EXISTS state(k TEXT PRIMARY KEY,v BLOB)'); c.commit(); return c
def save(k,v):
 c=dbconn(); c.execute('INSERT OR REPLACE INTO state(k,v) VALUES(?,?)',(k,sqlite3.Binary(pickle.dumps(v)))); c.commit(); c.close()
def load(k,d=None):
 c=dbconn(); r=c.execute('SELECT v FROM state WHERE k=?',(k,)).fetchone(); c.close()
 if not r:return d
 try:return pickle.loads(r[0])
 except:return d



# Firebase Admin / Firestore: server-side access using Streamlit Secrets.
def firebase_db():
    try:
        if not firebase_admin._apps:
            sa = dict(st.secrets.get('firebase_admin', {}))
            if not sa.get('project_id') or not sa.get('private_key') or not sa.get('client_email'):
                return None
            firebase_admin.initialize_app(credentials.Certificate(sa))
        return firestore.client()
    except Exception:
        return None

def _fs_delete_collection(db, name):
    refs = list(db.collection(name).stream())
    for i in range(0, len(refs), 450):
        batch = db.batch()
        for ref in refs[i:i+450]:
            batch.delete(ref.reference)
        batch.commit()

def _fs_save_df(db, name, df, key_col=None):
    if df is None:
        return
    _fs_delete_collection(db, name)
    records = json.loads(df.to_json(orient='records', date_format='iso'))
    batch = db.batch()
    pending = 0
    for idx, rec in enumerate(records):
        if key_col and key_col in rec:
            raw = str(rec.get(key_col) or '').strip()
            doc_id = raw if raw else f'row_{idx}'
        else:
            doc_id = hashlib.sha1(json.dumps(rec, sort_keys=True, ensure_ascii=False).encode('utf-8')).hexdigest()
        doc_id = doc_id.replace('/', '_')
        rec['_ordem'] = idx
        batch.set(db.collection(name).document(doc_id), rec)
        pending += 1
        if pending >= 450:
            batch.commit()
            batch = db.batch()
            pending = 0
    if pending:
        batch.commit()

def _fs_load_df(db, name):
    rows = [x.to_dict() for x in db.collection(name).stream()]
    if not rows:
        return None
    rows.sort(key=lambda x: x.get('_ordem', 0))
    for r in rows:
        r.pop('_ordem', None)
    return pd.DataFrame(rows)

def _fs_save_eligible(db, values):
    db.collection('estoque_config').document('enderecos').set({'enderecos': list(values or [])})

def _fs_load_eligible(db):
    snap = db.collection('estoque_config').document('enderecos').get()
    if not snap.exists:
        return None
    return list((snap.to_dict() or {}).get('enderecos') or [])

def firestore_load_db():
    db = firebase_db()
    if db is None:
        return None
    try:
        return _fs_load_df(db, 'estoque_produtos')
    except Exception:
        return None

def firestore_load_pos():
    db = firebase_db()
    if db is None:
        return None
    try:
        return _fs_load_df(db, 'estoque_posicoes')
    except Exception:
        return None

def firestore_load_eligible():
    db = firebase_db()
    if db is None:
        return None
    try:
        return _fs_load_eligible(db)
    except Exception:
        return None

def load_user_profile():
    user = st.session_state.get('auth_user') or {}
    uid = user.get('localId','')
    email = (user.get('email') or '').strip().lower()
    if not uid: return 'Operador'
    # Bootstrap admin must be recognized even if Firestore is temporarily unavailable.
    bootstrap = str(st.secrets.get('FIREBASE_BOOTSTRAP_ADMIN_EMAIL','')).strip().lower()
    if bootstrap and email and email == bootstrap:
        db = firebase_db()
        if db is not None:
            try:
                db.collection('usuarios').document(uid).set({'email':email,'nome':email.split('@')[0],'perfil':'ADMIN','ativo':True,'atualizado_em':firestore.SERVER_TIMESTAMP}, merge=True)
            except Exception:
                pass
        return 'Admin'
    db = firebase_db()
    if db is None: return 'Operador'
    try:
        ref=db.collection('usuarios').document(uid)
        snap=ref.get()
        if snap.exists:
            data=snap.to_dict() or {}
            if data.get('ativo') is False:
                auth_logout()
                return 'Operador'
            perfil=str(data.get('perfil','OPERADOR')).upper()
            return {'ADMIN':'Admin','GESTOR':'Gestor','OPERADOR':'Operador'}.get(perfil,'Operador')
        ref.set({'email':email,'nome':email.split('@')[0] if email else 'Usuário','perfil':'OPERADOR','ativo':True,'criado_em':firestore.SERVER_TIMESTAMP})
        return 'Operador'
    except Exception:
        return 'Operador'

if 'cfg' not in st.session_state: st.session_state.cfg={**DEFAULT,**(load('cfg',{}) or {})}
if 'logo' not in st.session_state: st.session_state.logo=load('logo',(None,''))
if 'db' not in st.session_state:
 _fsdb=firestore_load_db(); st.session_state.db=_fsdb if _fsdb is not None else load('db')
if 'pos' not in st.session_state:
 _fspos=firestore_load_pos(); st.session_state.pos=_fspos if _fspos is not None else load('pos')
if 'eligible' not in st.session_state:
 _fselig=firestore_load_eligible(); st.session_state.eligible=_fselig if _fselig is not None else (load('eligible',[]) or [])
if 'inventories' not in st.session_state: st.session_state.inventories=load('inventories',{}) or {}
if 'cycles' not in st.session_state: st.session_state.cycles=load('cycles',{}) or {}
if 'section' not in st.session_state: st.session_state.section='Dashboard'
if 'selected' not in st.session_state: st.session_state.selected=None
if 'new_inv' not in st.session_state: st.session_state.new_inv=False
if 'profile' not in st.session_state: st.session_state.profile='Operador'
if 'reports' not in st.session_state: st.session_state.reports=load('reports',{}) or {}

st.session_state.profile=load_user_profile()
cfg=st.session_state.cfg
config=cfg

# Compatibility defaults: preserve older saved settings while supporting the V5 UI keys.
_cfg_defaults = {
    'sidebar_subtitle':'SISTEMA OPERACIONAL DE ESTOQUE',
    'menu_label':'MENU',
    'dashboard_label':'DASHBOARD',
    'inventory_label':'INVENTÁRIO ROTATIVO',
    'database_label':'BANCO DE DADOS',
    'register_label':'REGISTRO',
    'settings_label':'CONFIGURAÇÕES',
    'report_label':'REPORTAR INCONSISTÊNCIAS',
    'new_inventory_text':'NOVO INVENTÁRIO',
    'address_title':'ENDEREÇOS ELEGÍVEIS',
    'sidebar_width':250,'menu_gap':2,
}
for _k, _v in _cfg_defaults.items():
    config.setdefault(_k, _v)
    cfg.setdefault(_k, _v)

# Authentication: Firebase Authentication (email + password). Session remains server-side in Streamlit.
FIREBASE_API_KEY = st.secrets.get('FIREBASE_API_KEY', 'AIzaSyDkS32UBjttYW1bWFho60EUnP4DXRYnKps')
FIREBASE_AUTH_URL = 'https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword'
if 'auth_user' not in st.session_state: st.session_state.auth_user=None

def auth_login(email, password):
    if not FIREBASE_API_KEY:
        return None, 'A autenticação ainda não foi configurada no aplicativo.'
    try:
        body=json.dumps({'email':email.strip().lower(),'password':password,'returnSecureToken':True}).encode('utf-8')
        req=Request(FIREBASE_AUTH_URL+'?key='+FIREBASE_API_KEY,data=body,headers={'Content-Type':'application/json'},method='POST')
        with urlopen(req,timeout=15) as resp: data=json.loads(resp.read().decode('utf-8'))
        return data, None
    except HTTPError as e:
        try:
            raw=e.read().decode('utf-8')
            info=json.loads(raw)
            code=(info.get('error') or {}).get('message','')
            messages={
                'INVALID_LOGIN_CREDENTIALS':'E-mail ou senha inválidos.',
                'INVALID_PASSWORD':'E-mail ou senha inválidos.',
                'EMAIL_NOT_FOUND':'E-mail ou senha inválidos.',
                'USER_DISABLED':'Este usuário está desativado.',
                'TOO_MANY_ATTEMPTS_TRY_LATER':'Muitas tentativas. Tente novamente mais tarde.'
            }
            msg=messages.get(code,'Não foi possível realizar o login.')
        except Exception:
            msg='Não foi possível realizar o login.'
        return None, msg
    except (URLError, TimeoutError):
        return None, 'Não foi possível conectar ao serviço de autenticação.'
    except Exception:
        return None, 'Não foi possível realizar o login.'

def auth_logout():
    st.session_state.auth_user=None
    st.session_state.pop('auth_access_token',None)
    st.session_state.pop('auth_refresh_token',None)
    st.rerun()

def render_login():
    b,n = st.session_state.logo
    login_logo = None
    if b:
        ext=n.lower()
        mime='image/png' if ext.endswith('.png') else 'image/jpeg' if ext.endswith(('.jpg','.jpeg')) else 'image/webp' if ext.endswith('.webp') else 'image/svg+xml'
        login_logo='data:'+mime+';base64,'+base64.b64encode(b).decode()
    st.markdown('''<style>
    [data-testid="stHeader"]{background:transparent!important}
    .login-wrap{min-height:0!important;display:flex;align-items:flex-start;justify-content:center;padding-top:12px;margin-bottom:24px}
    .login-card{width:min(440px,92vw);padding:18px 24px 14px;background:#101614;border:1px solid #2B3732;border-radius:18px;box-shadow:0 10px 30px rgba(0,0,0,.28);text-align:center}
    .login-logo{display:flex;align-items:center;justify-content:center;height:66px;margin:0 auto 4px;overflow:hidden}
    .login-logo img{max-width:250px;max-height:62px;width:auto;height:auto;object-fit:contain;display:block}
    .login-brand{text-align:center;font-size:34px;font-weight:900;letter-spacing:-1px;color:#F4F5F2;margin-bottom:4px}
    .login-brand span{color:#FFD63B}
    .login-sub{text-align:center;color:#A9B1AC;font-size:12px;margin-bottom:10px}
    .login-title{text-align:center;color:#F4F5F2;font-size:18px;font-weight:800;margin-bottom:0}
    [class*="st-key-login_form"]{margin-top:0!important}
    [class*="st-key-login_form"] [data-testid="stForm"]{border:1px solid #2B3732!important;border-radius:14px!important;padding:18px 20px 16px!important;background:transparent!important}
    [class*="st-key-login_form"] [data-testid="stFormSubmitButton"] button{margin-top:4px!important}
    .login-footer{width:min(440px,92vw);text-align:center;color:#A9B1AC;font-size:11px;margin:8px auto 0}
    </style>''',unsafe_allow_html=True)
    logo_html = (f'<div class="login-logo"><img src="{login_logo}"></div>' if login_logo else '<div class="login-brand">Se<span>tt</span>a</div>')
    st.markdown(f'<div class="login-wrap"><div class="login-card">{logo_html}<div class="login-sub">SISTEMA OPERACIONAL DE ESTOQUE</div><div class="login-title">ACESSO AO SISTEMA</div></div></div>',unsafe_allow_html=True)
    with st.form('login_form'):
        email=st.text_input('E-mail',placeholder='seu e-mail')
        password=st.text_input('Senha',type='password',placeholder='••••••••')
        submitted=st.form_submit_button('ENTRAR',type='primary',use_container_width=True,icon=':material/login:')
    if submitted:
        if not email or not password:
            st.error('Informe o e-mail e a senha.')
        else:
            data,err=auth_login(email,password)
            if err: st.error(err)
            else:
                # Firebase REST returns idToken/localId directly; there is no nested 'user' object.
                st.session_state.auth_user={'email': data.get('email', email), 'localId': data.get('localId','')}
                st.session_state.auth_access_token=data.get('idToken','')
                st.session_state.auth_refresh_token=data.get('refreshToken','')
                st.rerun()
    st.markdown('<div class="login-footer">Acesso autorizado somente para usuários cadastrados.</div>',unsafe_allow_html=True)

if not st.session_state.auth_user:
    render_login()
    st.stop()


def persist_cfg(): save('cfg',cfg)
def persist_all(): save('inventories',st.session_state.inventories); save('cycles',st.session_state.cycles)
def persist_eligible():
 db=firebase_db()
 if db is not None:
  try:
   _fs_save_eligible(db,st.session_state.eligible)
   return
  except Exception:
   pass
 save('eligible',st.session_state.eligible)

def persist_db():
 db=firebase_db()
 if db is not None:
  try:
   _fs_save_df(db,'estoque_produtos',st.session_state.db,'codigo')
   _fs_save_df(db,'estoque_posicoes',st.session_state.pos,None)
   _fs_save_eligible(db,st.session_state.eligible)
   return
  except Exception:
   pass
 save('db',st.session_state.db); save('pos',st.session_state.pos); save('eligible',st.session_state.eligible)
def persist_reports(): save('reports',st.session_state.reports)

def excel_bytes(df, sheet_name):
 out=io.BytesIO()
 with pd.ExcelWriter(out,engine='openpyxl') as writer:
  df.to_excel(writer,index=False,sheet_name=sheet_name)
  ws=writer.book[sheet_name]
  ws.freeze_panes='A2'
  ws.auto_filter.ref=ws.dimensions
  from openpyxl.styles import Font, PatternFill, Alignment
  for cell in ws[1]:
   cell.font=Font(bold=True,color='11130F')
   cell.fill=PatternFill('solid',fgColor='FFD63B')
   cell.alignment=Alignment(horizontal='center',vertical='center')
  ws.row_dimensions[1].height=24
  for col in ws.columns:
   letter=col[0].column_letter
   max_len=max(len(str(c.value)) if c.value is not None else 0 for c in col[:80])
   ws.column_dimensions[letter].width=min(max(max_len+2,10),60)
 return out.getvalue()

def signed_brl(v):
 x=float(v)
 if abs(x)<1e-12:return 'R$ 0,00'
 return ('+' if x>0 else '-')+brl(abs(x))

def logo_uri():
 b,n=st.session_state.logo
 if not b:return None
 ext=n.lower(); mime='image/png' if ext.endswith('.png') else 'image/jpeg' if ext.endswith(('.jpg','.jpeg')) else 'image/webp' if ext.endswith('.webp') else 'image/svg+xml'
 return 'data:'+mime+';base64,'+base64.b64encode(b).decode()

def css():
 d=cfg['theme']=='Dark'; vals=(cfg['dark_bg'],cfg['dark_panel'],cfg['dark_panel2'],cfg['dark_border'],cfg['dark_text'],cfg['dark_muted']) if d else (cfg['clean_bg'],cfg['clean_panel'],cfg['clean_panel2'],cfg['clean_border'],cfg['clean_text'],cfg['clean_muted']); bg,panel,panel2,border,text,muted=vals
 st.markdown(f'''<style>
:root{{--p:{cfg['primary']};--ph:{cfg['hover']};--bg:{bg};--panel:{panel};--p2:{panel2};--border:{border};--text:{text};--muted:{muted}}}
html,body,[class*="css"],.stApp{{font-family:{cfg['font']},Arial,sans-serif!important;font-size:{cfg['font_size']}px!important}}.stApp{{background:var(--bg);color:var(--text)}}[data-testid="stHeader"]{{background:var(--bg)}}
section[data-testid="stSidebar"]{{background:var(--bg);border-right:1px solid var(--border)}}section[data-testid="stSidebar"]>div{{padding-top:.25rem!important;position:relative!important}}section[data-testid="stSidebar"] button[aria-label*='Collapse'],section[data-testid="stSidebar"] button[aria-label*='Expand'],header button[aria-label*='Collapse'],header button[aria-label*='Expand']{{position:relative!important;z-index:99999!important;pointer-events:auto!important}}.logo-area{{pointer-events:none!important;position:relative;z-index:0}}
/* Controle nativo de abrir/recolher: não reposicionar */
.logo-area{{height:{cfg['logo_h']+25}px;display:flex;align-items:center;justify-content:{cfg['logo_align']};transform:translateY({cfg['logo_top']}px);padding:0 8px;overflow:hidden}}.logo-area img{{width:{cfg['logo_w']}px;height:{cfg['logo_h']}px;object-fit:contain;display:block}}
.sidebar-sub{{color:var(--muted);font-size:11px;text-align:{cfg['sidebar_align']};transform:translateY({cfg['sub_top']}px);margin:0 7px {max(4,cfg['gap'])}px}}.menu-label{{color:var(--p);font-weight:800;font-size:12px;margin:0 7px 8px;transform:translateY({cfg['menu_top']}px);text-align:{cfg['sidebar_align']}}}
section[data-testid="stSidebar"] .stButton{{width:100%!important;margin-bottom:{cfg.get('menu_gap',2)}px}}section[data-testid="stSidebar"] .stButton>button{{width:100%!important;min-height:{cfg['item_h']}px;height:{cfg['item_h']}px;border:1px solid var(--border);background:var(--panel);color:var(--text);text-align:left;justify-content:flex-start;font-size:{cfg['sidebar_font']}px;font-weight:800;border-radius:9px;padding:0 14px;box-shadow:0 1px 3px rgba(0,0,0,.18);gap:10px}}section[data-testid="stSidebar"] .stButton>button:hover{{background:var(--p2);color:var(--text);border-color:var(--p)}}section[data-testid="stSidebar"] .stButton>button[kind="primary"]{{background:var(--p);color:#11130F;border-color:var(--p);box-shadow:0 2px 7px rgba(0,0,0,.22)}}section[data-testid="stSidebar"] .stButton>button [data-testid="stIconMaterial"]{{font-size:20px!important;width:22px!important;min-width:22px!important;height:22px!important;line-height:22px!important;color:{cfg['icon_color']}!important;display:inline-flex!important;align-items:center!important;justify-content:center!important}}
.main-title{{font-size:{cfg['title_size']}px;font-weight:800;line-height:1.1;color:var(--text);margin:5px 0 2px}}.main-subtitle{{color:var(--muted);font-size:14px;margin-bottom:22px}}
[data-testid="stMetric"]{{background:var(--panel);border:1px solid var(--border);border-radius:14px;padding:17px 19px;min-height:105px}}[data-testid="stMetricLabel"] p{{color:var(--muted)!important;font-size:11px!important;font-weight:800!important;letter-spacing:.5px;text-transform:uppercase}}[data-testid="stMetricValue"]{{color:var(--text)}}
div[data-testid="stVerticalBlockBorderWrapper"]{{background:var(--panel);border-color:var(--border)!important;border-radius:14px}}.stButton>button,.stDownloadButton>button{{border-radius:8px;font-weight:800;border:1px solid var(--border);background:var(--p2);color:var(--text)}}.stButton>button:hover,.stDownloadButton>button:hover{{border-color:var(--p);color:var(--p)}}.stButton>button[kind="primary"]{{background:var(--p);color:#10120F;border-color:var(--p)}}label,.stMarkdown p,.stCaption,.stRadio label,.stCheckbox label{{color:var(--text)!important}}input,textarea{{color:var(--text)!important}}[data-testid="stDataFrame"]{{border:1px solid var(--border);border-radius:10px;overflow:hidden}}
section[data-testid="stSidebar"] .sidebar-report-spacer{{display:none!important}}section[data-testid="stSidebar"] .sidebar-report-area{{position:absolute;left:10px;right:10px;bottom:18px;margin:0;z-index:20}}section[data-testid="stSidebar"] .sidebar-report-area .stButton{{margin-bottom:0!important}}section[data-testid="stSidebar"] .sidebar-user-area{{position:absolute;left:10px;right:10px;bottom:78px;padding-top:6px;border-top:1px solid var(--border);z-index:19}}section[data-testid="stSidebar"] .sidebar-user-area .stCaption{{font-size:10px!important;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}section[data-testid="stSidebar"] .sidebar-user-area .stButton{{margin:0!important}}section[data-testid="stSidebar"] .sidebar-user-area .stButton>button{{min-height:32px;height:32px;font-size:11px!important}}</style>''',unsafe_allow_html=True)
css()

def ncode(s): return s.astype('string').fillna('').str.strip().str.replace(r'\.0$','',regex=True).str.zfill(8)
def naddr(s): return s.astype('string').fillna('').str.strip().str.replace(r'\s+',' ',regex=True).str.upper()
def readxls(f): f.seek(0); return pd.read_excel(f,sheet_name=0,header=1,dtype=str)
def pnum(v):
 if v is None or pd.isna(v):return 0.0
 s=str(v).strip().replace('R$','').replace(' ','')
 if not s:return 0.0
 neg=s.startswith('-');s=s.lstrip('+-')
 if ',' in s and '.' in s:s=s.replace('.','').replace(',','.') if s.rfind(',')>s.rfind('.') else s.replace(',','')
 elif ',' in s:s=s.replace(',','.')
 try:x=float(s);return -x if neg else x
 except:return 0.0
def nums(s):return s.apply(pnum).astype(float)
def fn(v,d=3):return f'{float(v):,.{d}f}'.replace(',','X').replace('.',',').replace('X','.')
def brl(v):return f'R$ {float(v):,.2f}'.replace(',','X').replace('.',',').replace('X','.')

def build_db(an,end,eligible):
 a=an.iloc[:,[0,3,7,10]].copy();a.columns=['codigo','descricao','qtd_analitico','valor_k'];a['codigo']=ncode(a.codigo);a['descricao']=a.descricao.astype('string').fillna('').str.strip();a['qtd_analitico']=nums(a.qtd_analitico);a['valor_k']=nums(a.valor_k)
 a=a.groupby('codigo',as_index=False).agg(descricao=('descricao','first'),qtd_analitico=('qtd_analitico','sum'),valor_k=('valor_k','sum'));a['valor_unitario']=a.apply(lambda r:r.valor_k/r.qtd_analitico if abs(r.qtd_analitico)>1e-12 else 0,axis=1)
 e=end.iloc[:,[0,3,7]].copy();e.columns=['codigo','endereco','quantidade'];e['codigo']=ncode(e.codigo);e['endereco']=naddr(e.endereco);e['quantidade']=nums(e.quantidade);e=e.groupby(['codigo','endereco'],as_index=False).quantidade.sum(); ap=set(naddr(pd.Series(eligible)).tolist());e['apto']=e.endereco.isin(ap)
 saldo=e[e.apto].groupby('codigo',as_index=False).quantidade.sum().rename(columns={'quantidade':'saldo_apto'});d=a.merge(saldo,on='codigo',how='outer');d.saldo_apto=d.saldo_apto.fillna(0.0);d.qtd_analitico=d.qtd_analitico.fillna(0.0);d.valor_k=d.valor_k.fillna(0.0);d.valor_unitario=d.valor_unitario.fillna(0.0);d.descricao=d.descricao.fillna('SEM DESCRIÇÃO NO ESTOQUE ANALÍTICO');d['valor_total']=d.saldo_apto*d.valor_unitario;act=(d.saldo_apto>0)&(d.valor_unitario>0);d['classificacao_r_un']=pd.NA;d['classificacao_r_total']=pd.NA;d.loc[act,'classificacao_r_un']=d.loc[act].valor_unitario.rank(method='first',ascending=False).astype(int);d.loc[act,'classificacao_r_total']=d.loc[act].valor_k.rank(method='first',ascending=False).astype(int);d=d.sort_values(['classificacao_r_total','codigo'],na_position='last').reset_index(drop=True);e=e.merge(d[['codigo','valor_unitario']],on='codigo',how='left');return d,e

def nextdoc():
 today=datetime.now().strftime('%d%m%Y');q=1
 for x in st.session_state.inventories.values():
  if x.get('documento','').startswith(today+'-'):
   try:q=max(q,int(x['documento'].split('-')[-1])+1)
   except:pass
 return f'{today}-{q:03d}'
def cycle():
 codes=st.session_state.db.loc[st.session_state.db.saldo_apto>0,'codigo'].astype(str);return min([int(st.session_state.cycles.get(c,0)) for c in codes],default=0)+1
def select_products(db,n,urgent_codes=None):
 w=db[(db.saldo_apto>0)&(db.valor_unitario>0)].copy();w['cc']=w.codigo.astype(str).map(lambda c:int(st.session_state.cycles.get(c,0)));m=w.cc.min();w=w[w.cc==m];u=w.sort_values(['classificacao_r_un','codigo'],na_position='last');t=w.sort_values(['classificacao_r_total','codigo'],na_position='last');nu=n//2;sel=[]
 for c in u.codigo:
  if len(sel)>=nu:break
  if c not in sel:sel.append(c)
 for c in t.codigo:
  if len(sel)>=n:break
  if c not in sel:sel.append(c)
 urgent_codes=[str(c) for c in (urgent_codes or [])]
 code_set=set(db.codigo.astype(str))
 urgent_available=[c for c in urgent_codes if c in code_set and c not in sel]
 for c in urgent_available:
  sel.append(c)
 target=n+len(urgent_available)
 if len(sel)<target:
  for c in pd.concat([u,t]).drop_duplicates('codigo').codigo:
   c=str(c)
   if c not in sel:sel.append(c)
   if len(sel)>=target:break
 return db[db.codigo.astype(str).isin(sel)].copy()
def make_rows(sel,pos):
 rows=[]
 for _,p in sel.iterrows():
  for _,r in pos[(pos.codigo==p.codigo)&pos.apto].iterrows():
   rows.append({'id':uuid.uuid4().hex[:12],'codigo':str(p.codigo),'descricao':str(p.descricao),'endereco':str(r.endereco),'qtd_sistema':float(r.quantidade),'valor_unitario':float(p.valor_unitario),'contagens':[],'status':'PENDENTE','contagem_final':None,'resultado_final':'','comentario_final':'SC'})
 return rows
def persist_inv(inv):st.session_state.inventories[inv['documento']]=inv;save('inventories',st.session_state.inventories)
def addcount(r,q,cm,stage):r['contagens'].append({'etapa':stage,'quantidade':float(q),'comentario':cm.strip() if cm.strip() else 'SC','data':datetime.now().strftime('%d/%m/%Y %H:%M:%S')})
def last(r):return r['contagens'][-1]['quantidade'] if r['contagens'] else None
def diff(r,q):return float(q)-float(r['qtd_sistema'])
def divergencia_valor(r,q):return diff(r,q)*float(r['valor_unitario'])
def divergência(r,q):return abs(divergencia_valor(r,q))
def sev(v):return 'BAIXO' if v<=100 else 'MÉDIO' if v<=1000 else 'ALTO'
def mark_cycle(inv):
 if inv.get('ciclo_marcado'):return
 for c in {r['codigo'] for r in inv['rows'] if r['contagens']}:st.session_state.cycles[c]=int(st.session_state.cycles.get(c,0))+1
 inv['ciclo_marcado']=True;save('cycles',st.session_state.cycles);persist_inv(inv)
def close_inv(inv):
 for r in inv['rows']:
  if r['contagem_final'] is None:r['contagem_final']=last(r)
  if not r['resultado_final']:r['resultado_final']='ENCERRADO PELO GESTOR'
  r['status']='FINALIZADO'
 inv['status']='FECHADO';mark_cycle(inv)
 for rep in st.session_state.reports.values():
  if rep.get('status')=='ABERTO' and any(str(r['codigo'])==str(rep.get('codigo')) and str(r['endereco'])==str(rep.get('endereco')) for r in inv['rows']):
   rep['status']='ENCERRADO'
   rep['inventario_doc']=inv['documento']
   rep['encerrado_em']=datetime.now().strftime('%d/%m/%Y %H:%M:%S')
 persist_reports();persist_inv(inv)

# Sidebar
with st.sidebar:
 u=logo_uri()
 if u:st.markdown(f'<div class="logo-area"><img src="{u}"></div>',unsafe_allow_html=True)
 else:st.markdown('<div class="logo-area"><div style="color:var(--muted);text-align:center;font-size:11px">LOGO DA EMPRESA<br>Configure em Configurações.</div></div>',unsafe_allow_html=True)
 st.markdown(f'<div class="sidebar-sub">{config["sidebar_subtitle"]}</div>',unsafe_allow_html=True);st.markdown(f'<div class="menu-label">{config["menu_label"]}</div>',unsafe_allow_html=True)
 nav=[('Dashboard',config["dashboard_label"],':material/dashboard:'),('Inventário Rotativo',config["inventory_label"],':material/inventory_2:'),('Banco de Dados',config["database_label"],':material/database:'),('Registro',config["register_label"],':material/history:'),('Configurações',config["settings_label"],':material/settings:')]
 if st.session_state.profile=='Admin': nav.insert(4,('Usuários','USUÁRIOS',':material/manage_accounts:'))
 tops={'Dashboard':cfg['dash_top'],'Inventário Rotativo':cfg['inv_top'],'Banco de Dados':cfg['db_top'],'Registro':cfg['reg_top'],'Configurações':cfg['settings_top'],'Usuários':0}
 for k,l,ic in nav:
  off=tops.get(k,0)
  st.markdown(f'<div style="height:0;margin-top:{off}px"></div>',unsafe_allow_html=True)
  if st.button(l,key='nav_'+k,icon=ic,icon_position='left',type='primary' if st.session_state.section==k else 'secondary'):st.session_state.section=k;st.rerun()
 st.markdown('<div class="sidebar-report-area">',unsafe_allow_html=True)
 if st.button(config["report_label"],key='nav_Reportar Inconsistências',icon=':material/report_problem:',icon_position='left',type='primary' if st.session_state.section=='Reportar Inconsistências' else 'secondary'):
  st.session_state.section='Reportar Inconsistências';st.rerun()
 st.markdown('</div>',unsafe_allow_html=True)
 st.markdown('<div class="sidebar-user-area">',unsafe_allow_html=True)
 _uemail=str(st.session_state.auth_user.get('email',''))
 st.caption('ACESSO: '+_uemail)
 if st.button('SAIR',key='logout_btn',icon=':material/logout:',use_container_width=True): auth_logout()
 st.markdown('</div>',unsafe_allow_html=True)

st.markdown(f'<div class="main-title">{config["title"]}</div><div class="main-subtitle">{config["subtitle"]}</div>',unsafe_allow_html=True)
active=st.session_state.section

# Dashboard
if active=='Dashboard':
 st.subheader(config['dashboard_title']);st.caption('Visão geral dos indicadores do estoque.')
 if st.session_state.db is None:st.info('Importe e processe os relatórios na aba Banco de Dados.')
 else:
  db=st.session_state.db;items=int((db.saldo_apto>0).sum());valor_apto=float(db.valor_total.sum());rr=[r for x in st.session_state.inventories.values() for r in x['rows']];cnt=[r for r in rr if r['contagens']];div=[r for r in cnt if abs(diff(r,last(r)))>1e-9]
  qtd_cnt=len(cnt);qtd_div=len(div);acc_itens=(100-(qtd_div/items*100)) if items else 100.0;acc_pos=(100-(qtd_div/qtd_cnt*100)) if qtd_cnt else 100.0
  a,b,c,d=st.columns(4);a.metric('ITENS DIFERENTES COM SALDO',f'{items:,}'.replace(',','.'));b.metric('VALOR TOTAL APTO A CONTABILIZAR',brl(valor_apto));c.metric('POSIÇÕES CONTABILIZADAS',f'{qtd_cnt:,}'.replace(',','.'));d.metric('POSIÇÕES DIVERGENTES',f'{qtd_div:,}'.replace(',','.'))
  a,b=st.columns(2);a.metric('ACURÁCIA · DIVERGENTES / ITENS COM SALDO',f'{acc_itens:.2f}%');b.metric('ACURÁCIA · DIVERGENTES / CONTABILIZADOS',f'{acc_pos:.2f}%')
  st.markdown('#### Indicadores visuais')
  ch1,ch2=st.columns(2)
  with ch1:
   import altair as alt
   status_df=pd.DataFrame({'Status':['Sem divergência','Com divergência'],'Quantidade':[max(qtd_cnt-qtd_div,0),qtd_div]})
   chart1=alt.Chart(status_df).mark_bar(cornerRadiusTopLeft=7,cornerRadiusTopRight=7,size=72).encode(
    x=alt.X('Status:N',sort=['Sem divergência','Com divergência'],axis=alt.Axis(title=None,labelAngle=0)),
    y=alt.Y('Quantidade:Q',axis=alt.Axis(title=None,grid=True,gridColor='#FFFFFF',gridOpacity=0.28,tickColor='#FFFFFF',labelColor='#FFFFFF')),
    color=alt.value('#FFD63B'),
    tooltip=[alt.Tooltip('Status:N',title='Status'),alt.Tooltip('Quantidade:Q',title='Posições')]
   ).properties(height=260,background='transparent')
   st.altair_chart(chart1,use_container_width=True)
  with ch2:
   inv_rows=[]
   for x in sorted(st.session_state.inventories.values(),key=lambda z:z.get('criado_em','')):
    total=sum(1 for r in x['rows'] if r['contagens']);dv=sum(1 for r in x['rows'] if r['contagens'] and abs(diff(r,last(r)))>1e-9)
    if total:inv_rows.extend([{'Inventário':x['documento'],'Status':'Contabilizadas','Quantidade':total},{'Inventário':x['documento'],'Status':'Divergentes','Quantidade':dv}])
   if inv_rows:
    chart=pd.DataFrame(inv_rows)
    chart2=alt.Chart(chart).mark_bar(cornerRadiusTopLeft=5,cornerRadiusTopRight=5,size=26).encode(
     x=alt.X('Inventário:N',axis=alt.Axis(title=None,labelAngle=-45)),
     y=alt.Y('Quantidade:Q',axis=alt.Axis(title=None,grid=True,gridColor='#FFFFFF',gridOpacity=0.28,tickColor='#FFFFFF',labelColor='#FFFFFF')),
     xOffset=alt.XOffset('Status:N'),
     color=alt.value('#FFD63B'),
     tooltip=[alt.Tooltip('Inventário:N',title='Inventário'),alt.Tooltip('Status:N',title='Status'),alt.Tooltip('Quantidade:Q',title='Posições')]
    ).properties(height=260,background='transparent')
    st.altair_chart(chart2,use_container_width=True)
   else:
    st.info('Ainda não existem contagens para gerar o gráfico por inventário.')

# Inventory
elif active=='Inventário Rotativo':
 st.subheader(config['inventory_title']);st.caption(config['inventory_subtitle'])
 if st.session_state.db is None:st.info('Primeiro importe e processe a base na aba Banco de Dados.')
 else:
  a,b=st.columns(2);st.session_state.profile=a.radio('Perfil de teste',['Operador','Gestor'],index=0 if st.session_state.profile=='Operador' else 1,horizontal=True)
  if b.button(config['new_inventory_text'],type='primary',use_container_width=True):st.session_state.new_inv=True;st.rerun()
  if st.session_state.new_inv:
   with st.container(border=True):
    a,b,c=st.columns(3);n=a.number_input('Quantidade de produtos distintos',1,500,10);blind=b.checkbox('Contagem cega',value=cfg['blind_default']);c.metric('Ciclo atual',cycle());x,y=st.columns(2)
    if x.button('Criar inventário',type='primary',use_container_width=True):
     urgent_codes=sorted({str(rep.get('codigo')) for rep in st.session_state.reports.values() if rep.get('status')=='ABERTO' and rep.get('equipe')=='INVENTÁRIO ROTATIVO'})
     sel=select_products(st.session_state.db,n,urgent_codes);rows=make_rows(sel,st.session_state.pos)
     if not rows:st.error('Os produtos selecionados não possuem endereços aptos.')
     else:
      doc=nextdoc();st.session_state.inventories[doc]={'documento':doc,'data':datetime.now().strftime('%d/%m/%Y %H:%M'),'responsavel':st.session_state.profile,'blind_count':blind,'ciclo':cycle(),'status':'EM CONTAGEM','rows':rows,'criado_em':datetime.now().isoformat(timespec='seconds'),'ciclo_marcado':False};persist_inv(st.session_state.inventories[doc]);st.session_state.selected=doc;st.session_state.new_inv=False;st.rerun()
    if y.button('Cancelar'):st.session_state.new_inv=False;st.rerun()
  def render_inventory_cards(items):
   for inv in items:
    with st.container(border=True):
     s=len(inv['rows']);prod=len({r['codigo'] for r in inv['rows']});a,b,c,d=st.columns([2.2,1.5,1,1]);a.markdown(f'**{inv["documento"]}**');a.caption(f'Ciclo {inv["ciclo"]} · {inv["data"]}');b.write(f'**{inv["status"]}**');c.metric('Produtos',prod);d.metric('Posições',s)
     if st.button('Abrir',key='op_'+inv['documento']):st.session_state.selected=inv['documento'];st.rerun()
  all_inv=sorted(st.session_state.inventories.values(),key=lambda x:x.get('criado_em',''),reverse=True)
  open_inv=[x for x in all_inv if x.get('status')!='FECHADO']
  closed_inv=[x for x in all_inv if x.get('status')=='FECHADO']
  tab_open,tab_closed=st.tabs([f'EM ABERTO ({len(open_inv)})',f'FECHADOS ({len(closed_inv)})'])
  with tab_open:
   if open_inv:render_inventory_cards(open_inv)
   else:st.info('Nenhum inventário em aberto.')
  with tab_closed:
   if closed_inv:render_inventory_cards(closed_inv)
   else:st.info('Nenhum inventário fechado.')
  doc=st.session_state.selected
  if doc in st.session_state.inventories:
   inv=st.session_state.inventories[doc];st.divider();st.markdown(f'### Inventário {doc} — {inv["status"]}')
   prof=st.session_state.profile
   if prof=='Operador' and inv['status']=='EM CONTAGEM':
    for r in inv['rows']:
     if r['contagens']:continue
     with st.container(border=True):
      a,b,c=st.columns([1.1,3,1.3]);a.markdown(f'**{r["codigo"]}**');b.write(f'{r["descricao"]}\n\n**Endereço:** {r["endereco"]}');c.write('**Qtd. sistema:** OCULTA' if inv['blind_count'] else f'**Qtd. sistema:** {fn(r["qtd_sistema"])}');x,y=st.columns([1,2]);q=x.number_input('Contagem',0.0,step=.001,format='%.3f',key='q1_'+r['id']);cm=y.text_input('Comentário (opcional)',key='cm1_'+r['id'])
      if st.button('Salvar contagem',key='sv1_'+r['id'],type='primary'):addcount(r,q,cm,'1ª CONTAGEM');persist_inv(inv);st.rerun()
    if all(r['contagens'] for r in inv['rows']):
     if st.button('Fechar Contagem',type='primary'):inv['status']='AGUARDANDO ANÁLISE';persist_inv(inv);st.rerun()
   elif prof=='Gestor' and inv['status']=='AGUARDANDO ANÁLISE':
    st.markdown('#### Análise da 1ª contagem')
    # 1ª contagem: igual ao sistema confirma; divergente segue para decisão.
    for r in inv['rows']:
     if not r['contagens'] or r['status']=='FINALIZADO':
      continue
     q=last(r)
     if abs(diff(r,q))<1e-9:
      r['contagem_final']=q;r['resultado_final']='SISTEMA CONFIRMADO';r['status']='FINALIZADO'
    persist_inv(inv)
    divs=[r for r in inv['rows'] if r['contagens'] and r['status']!='FINALIZADO' and abs(diff(r,last(r)))>1e-9]
    if not divs: st.success('Não existem divergências na 1ª contagem. Todos os itens foram confirmados pelo sistema.')
    for r in divs:
     with st.container(border=True):
      q=last(r);a,b,c,d=st.columns(4);a.markdown(f'**{r["codigo"]} / {r["endereco"]}**');b.metric('Sistema',fn(r['qtd_sistema']));c.metric('1ª contagem',fn(q));d.metric('Divergência',signed_brl(divergencia_valor(r,q)));st.caption(f'Divergência de quantidade: {diff(r,q):+,.3f}'.replace(',','X').replace('.',',').replace('X','.').replace('+','+')+f' · Classificação: {sev(abs(divergencia_valor(r,q)))} · Comentário: {r["contagens"][-1]["comentario"]}')
      x,y,z=st.columns(3)
      if x.button('RECONTAR ESTE ITEM',key='r1_'+r['id']):r['status']='RECONTAR';inv['status']='AGUARDANDO RECONTAGEM';persist_inv(inv);st.rerun()
      if y.button('AUDITAR ESTE ITEM',key='a1_'+r['id']):r['status']='AUDITORIA';inv['status']='AGUARDANDO AUDITORIA';persist_inv(inv);st.rerun()
      if z.button('ENCERRAR ESTE ITEM',key='e1_'+r['id']):r['contagem_final']=q;r['resultado_final']='ENCERRADO PELO GESTOR';r['status']='FINALIZADO';persist_inv(inv);st.rerun()
    x,y,z=st.columns(3)
    if x.button('RECONTAR TODOS OS DIVERGENTES',type='primary',use_container_width=True):
     for r in inv['rows']:
      if r['contagens'] and r['status']!='FINALIZADO' and abs(diff(r,last(r)))>1e-9:r['status']='RECONTAR'
     inv['status']='AGUARDANDO RECONTAGEM';persist_inv(inv);st.rerun()
    if y.button('AUDITAR TODOS OS DIVERGENTES',use_container_width=True):
     for r in inv['rows']:
      if r['contagens'] and r['status']!='FINALIZADO' and abs(diff(r,last(r)))>1e-9:r['status']='AUDITORIA'
     inv['status']='AGUARDANDO AUDITORIA';persist_inv(inv);st.rerun()
    if z.button('ENCERRAR INVENTÁRIO',type='primary',use_container_width=True):close_inv(inv);st.rerun()
   elif prof=='Operador' and inv['status']=='AGUARDANDO RECONTAGEM':
    st.markdown('#### Recontagem — itens liberados pelo gestor')
    targets=[r for r in inv['rows'] if r['status']=='RECONTAR']
    for r in targets:
     with st.container(border=True):
      st.markdown(f'**{r["codigo"]} / {r["endereco"]}**');st.write(r['descricao']);st.caption('Histórico: '+' → '.join(f"{x['etapa']}: {fn(x['quantidade'])}" for x in r['contagens']));q=st.number_input('Nova contagem',0.0,step=.001,format='%.3f',key='q2_'+r['id']);cm=st.text_input('Comentário (opcional)',key='cm2_'+r['id'])
      if st.button('Salvar nova contagem',key='sv2_'+r['id'],type='primary'):addcount(r,q,cm,f'{len(r["contagens"])+1}ª CONTAGEM');r['status']='RECONTADA';persist_inv(inv);st.rerun()
    if not any(r['status']=='RECONTAR' for r in inv['rows']):inv['status']='AGUARDANDO DECISÃO';persist_inv(inv);st.rerun()
   elif prof=='Gestor' and inv['status']=='AGUARDANDO DECISÃO':
    st.markdown('#### Avaliação após cada recontagem')
    # Mesma regra para 2ª, 3ª e todas as contagens seguintes.
    candidates=[r for r in inv['rows'] if r['contagens'] and r['status']!='FINALIZADO' and abs(diff(r,last(r)))>1e-9]
    if not candidates: st.success('Não existem itens pendentes de decisão.')
    for r in candidates:
     q2=last(r);q1=r['contagens'][-2]['quantidade'] if len(r['contagens'])>=2 else None
     with st.container(border=True):
      st.markdown(f'**{r["codigo"]} / {r["endereco"]}**');st.write(f'Sistema: **{fn(r["qtd_sistema"])}** · Anterior: **{fn(q1) if q1 is not None else "—"}** · Atual: **{fn(q2)}**')
      if q1 is not None and abs(q2-q1)<1e-9:
       st.success('Atual igual à anterior: ERRO DE INVENTÁRIO. Nenhuma nova contagem é necessária.');r['contagem_final']=q2;r['resultado_final']='ERRO DE INVENTÁRIO';r['status']='FINALIZADO';persist_inv(inv)
      elif abs(q2-r['qtd_sistema'])<1e-9:
       st.success('Atual igual ao sistema: SISTEMA CONFIRMADO. Nenhuma nova contagem é necessária.');r['contagem_final']=q2;r['resultado_final']='SISTEMA CONFIRMADO';r['status']='FINALIZADO';persist_inv(inv)
      else:
       st.warning('A divergência permanece. O gestor deve decidir o próximo passo.');a,b,c=st.columns(3)
       if a.button('RECONTAR ESTE ITEM',key='dr_'+r['id']):r['status']='RECONTAR';inv['status']='AGUARDANDO RECONTAGEM';persist_inv(inv);st.rerun()
       if b.button('AUDITAR ESTE ITEM',key='da_'+r['id']):r['status']='AUDITORIA';inv['status']='AGUARDANDO AUDITORIA';persist_inv(inv);st.rerun()
       if c.button('ENCERRAR ESTE ITEM',key='dc_'+r['id']):r['contagem_final']=q2;r['resultado_final']='ENCERRADO PELO GESTOR';r['status']='FINALIZADO';persist_inv(inv);st.rerun()
    x,y,z=st.columns(3)
    if x.button('RECONTAR TODOS OS DIVERGENTES',type='primary',use_container_width=True):
     for r in inv['rows']:
      if r['contagens'] and r['status']!='FINALIZADO' and abs(diff(r,last(r)))>1e-9:r['status']='RECONTAR'
     inv['status']='AGUARDANDO RECONTAGEM';persist_inv(inv);st.rerun()
    if y.button('AUDITAR TODOS OS DIVERGENTES',use_container_width=True):
     for r in inv['rows']:
      if r['contagens'] and r['status']!='FINALIZADO' and abs(diff(r,last(r)))>1e-9:r['status']='AUDITORIA'
     inv['status']='AGUARDANDO AUDITORIA';persist_inv(inv);st.rerun()
    if z.button('ENCERRAR INVENTÁRIO',type='primary',use_container_width=True):close_inv(inv);st.rerun()
   elif prof=='Gestor' and inv['status']=='AGUARDANDO AUDITORIA':
    st.markdown('#### Auditoria / 3ª ou próxima contagem')
    targets=[r for r in inv['rows'] if r['status']=='AUDITORIA']
    for r in targets:
     with st.container(border=True):
      st.markdown(f'**{r["codigo"]} / {r["endereco"]}**');st.write(r['descricao']);st.caption('Histórico: '+' → '.join(f"{x['etapa']}: {fn(x['quantidade'])}" for x in r['contagens']));q=st.number_input('Contagem de auditoria',0.0,step=.001,format='%.3f',key='q3_'+r['id']);cm=st.text_input('Comentário (opcional)',key='cm3_'+r['id'])
      if st.button('Salvar auditoria',key='sv3_'+r['id'],type='primary'):addcount(r,q,cm,'AUDITORIA');r['status']='AUDITADA';persist_inv(inv);st.rerun()
    if not any(r['status']=='AUDITORIA' for r in inv['rows']):inv['status']='AGUARDANDO DECISÃO';persist_inv(inv);st.rerun()
   elif prof=='Gestor' and inv['status']=='FECHADO':
    st.success('Inventário encerrado e salvo no registro.');
    if st.button('Reabrir análise'):inv['status']='AGUARDANDO DECISÃO';persist_inv(inv);st.rerun()

# Database
elif active=='Banco de Dados':
 st.subheader(config['database_title']);st.caption('Base atual: ESTOQUE ANALÍTICO + ENDEREÇO. Lote continua ignorado.')
 a,b=st.columns(2)
 with a:
  f=st.file_uploader('1. Relatório ESTOQUE ANALÍTICO',type=['xlsx','xlsm','xltx'],key='up_an',help='A=Código, D=Descrição, H=Quantidade/Saldo, K=Valor em Estoque')
  if f:st.session_state.an_df=readxls(f);save('an_name',f.name);st.success(f'Carregado: {f.name} · {len(st.session_state.an_df):,} linhas')
 with b:
  f=st.file_uploader('2. Relatório ENDEREÇO',type=['xlsx','xlsm','xltx'],key='up_en',help='A=Código, D=Endereço, H=Quantidade')
  if f:st.session_state.en_df=readxls(f);save('en_name',f.name);st.success(f'Carregado: {f.name} · {len(st.session_state.en_df):,} linhas')
 if 'an_df' not in st.session_state:st.session_state.an_df=None
 if 'en_df' not in st.session_state:st.session_state.en_df=None
 if st.session_state.an_df is not None and st.session_state.en_df is not None:
  st.divider();st.subheader(config['address_title']);addresses=sorted([x for x in naddr(st.session_state.en_df.iloc[:,3]).unique() if x])
  if not st.session_state.eligible:st.session_state.eligible=addresses.copy()
  q=st.text_input('Pesquisar endereço',placeholder='Ex.: G9-M3-A-C1');shown=[x for x in addresses if q.strip().upper() in x] if q.strip() else addresses
  a,b,c=st.columns(3)
  if a.button('Marcar exibidos'):st.session_state.eligible=sorted(set(st.session_state.eligible)|set(shown));persist_db();st.rerun()
  if b.button('Desmarcar exibidos'):st.session_state.eligible=[x for x in st.session_state.eligible if x not in set(shown)];persist_db();st.rerun()
  if c.button('Marcar todos'):st.session_state.eligible=addresses.copy();persist_db();st.rerun()
  selected=set(st.session_state.eligible);st.caption(f'{len(shown)} endereços exibidos · {len(selected)} aptos');cols=st.columns(4)
  for i,addr in enumerate(shown):
   with cols[i%4]:
    v=st.checkbox(addr,value=addr in selected,key='address_'+str(abs(hash(addr))))
    if v!=(addr in selected):
     selected.add(addr) if v else selected.discard(addr);st.session_state.eligible=sorted(selected);persist_eligible()
  a,b=st.columns(2);a.metric('Endereços encontrados',len(addresses));b.metric('Endereços aptos',len(st.session_state.eligible))
  if st.button('PROCESSAR E ATUALIZAR BANCO',type='primary'):
   try:
    d,pos=build_db(st.session_state.an_df,st.session_state.en_df,st.session_state.eligible);st.session_state.db=d;st.session_state.pos=pos;st.session_state.cycles={};persist_db();save('cycles',{});st.success('Banco processado e salvo.')
   except Exception as e:st.error(f'Erro: {e}')
 if st.session_state.db is not None:
  st.divider();st.subheader('Banco consolidado');v=st.session_state.db.copy();v['valor_unitario']=v.valor_unitario.map(brl);v['saldo_apto']=v.saldo_apto.map(fn);v['valor_k']=v.valor_k.map(brl);v['valor_total']=v.valor_total.map(brl);v.columns=['Código','Descrição','Qtd. Analítico','Valor Total K','Valor Unitário','Saldo Apto','Valor Total Apto','Classificação R$ UN.','Classificação R$ TOTAL'];st.dataframe(v,use_container_width=True,hide_index=True,height=500);st.download_button('Exportar banco em Excel',excel_bytes(v,'Banco Consolidado'),'banco_consolidado.xlsx','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet');st.caption('Valor Unitário = K ÷ H. Valor Total do saldo apto = Saldo Apto × Valor Unitário. A Classificação R$ TOTAL usa o valor K do Estoque Analítico.')

# Register
elif active=='Registro':
 st.subheader(config['register_title']);st.caption(config['register_subtitle']);rows=[]
 for inv in st.session_state.inventories.values():
  if inv['status']!='FECHADO':continue
  for r in inv['rows']:
   rows.append({'Documento':inv['documento'],'Data':inv['data'],'Responsável':inv['responsavel'],'Ciclo':inv['ciclo'],'Código':r['codigo'],'Descrição':r['descricao'],'Endereço':r['endereco'],'Qtd. Sistema':r['qtd_sistema'],'Contagens':' | '.join(f"{x['etapa']}: {fn(x['quantidade'])} ({x['comentario']})" for x in r['contagens']),'Contagem Final':r['contagem_final'],'Resultado':r['resultado_final'],'Valor Divergência':divergencia_valor(r,r['contagem_final']) if r['contagem_final'] is not None else 0})
 if rows:
  df=pd.DataFrame(rows);display_df=df.copy();display_df['Valor Divergência']=display_df['Valor Divergência'].map(signed_brl);st.dataframe(display_df,use_container_width=True,hide_index=True);export_df=display_df.copy();st.download_button('Exportar Registro em Excel',excel_bytes(export_df,'Registro'),'registro_inventarios.xlsx','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
 else:st.info('Nenhum inventário fechado.')

# Reportar Inconsistências
elif active=='Reportar Inconsistências':
 st.subheader('Reportar Inconsistências');st.caption('Abertura e acompanhamento de inconsistências operacionais. O reporte de estoque será direcionado ao Inventário Rotativo.')
 if st.session_state.db is None:
  st.warning('Primeiro importe e processe a base na aba Banco de Dados.')
 else:
  if st.button('INICIAR INCONSISTÊNCIA',type='primary',use_container_width=True):st.session_state.new_report=True;st.rerun()
  if 'new_report' not in st.session_state:st.session_state.new_report=False
  if st.session_state.new_report:
   with st.container(border=True):
    st.markdown('#### Novo chamado de inconsistência')
    st.info('O material reportado será incluído automaticamente no próximo Inventário Rotativo.')
    codes=sorted(st.session_state.db['codigo'].astype(str).unique())
    addresses=sorted([str(x) for x in st.session_state.eligible if str(x).strip()])
    a,b=st.columns(2)
    equipe=a.selectbox('Equipe responsável',['INVENTÁRIO ROTATIVO'])
    codigo=b.selectbox('Código do material',options=['']+codes,index=0,help='Digite para pesquisar; a lista sugere correspondências.')
    descricao=''
    if codigo:
     m=st.session_state.db[st.session_state.db.codigo.astype(str)==str(codigo)]
     if not m.empty:descricao=str(m.iloc[0]['descricao'])
    st.caption(f'Descrição: {descricao}' if descricao else 'Selecione o código do material.')
    endereco=st.selectbox('Endereço',options=['']+addresses,index=0,help='Digite para pesquisar entre os endereços habilitados no Banco de Dados.')
    obs=st.text_area('OBSERVAÇÃO OBRIGATÓRIA',placeholder='INFORME O QUE ACONTECEU...',height=130)
    st.caption('A observação será registrada automaticamente em CAIXA ALTA.')
    x,y=st.columns(2)
    if x.button('SALVAR INCONSISTÊNCIA',type='primary',use_container_width=True):
     obs=obs.strip().upper()
     if not codigo or not endereco or not obs:
      st.error('Código, endereço e observação são obrigatórios.')
     else:
      rid=datetime.now().strftime('%Y%m%d%H%M%S')+'-'+uuid.uuid4().hex[:6].upper()
      st.session_state.reports[rid]={'id':rid,'criado_em':datetime.now().strftime('%d/%m/%Y %H:%M:%S'),'equipe':equipe,'codigo':str(codigo),'descricao':descricao,'endereco':str(endereco),'observacao':obs,'status':'ABERTO','inventario_doc':None,'encerrado_em':None}
      persist_reports();st.session_state.new_report=False;st.success(f'Inconsistência {rid} registrada.');st.rerun()
    if y.button('CANCELAR',use_container_width=True):st.session_state.new_report=False;st.rerun()
  reports=sorted(st.session_state.reports.values(),key=lambda x:x.get('criado_em',''),reverse=True)
  abertos=[r for r in reports if r.get('status')=='ABERTO'];encerrados=[r for r in reports if r.get('status')=='ENCERRADO']
  st.divider();st.markdown('### Chamados')
  ta,te=st.tabs([f'NÃO TRATADOS ({len(abertos)})',f'ENCERRADOS ({len(encerrados)})'])
  with ta:
   if not abertos:st.success('Nenhuma inconsistência pendente.')
   for r in abertos:
    with st.container(border=True):
     a,b,c=st.columns([1.2,2.2,1.5]);a.markdown(f'**{r["id"]}**');b.markdown(f'**{r["codigo"]}** · {r["descricao"]}');c.markdown(f'**{r["endereco"]}**')
     st.caption(f'Criado em {r["criado_em"]} · Equipe: {r["equipe"]}')
     st.write(f'**OBSERVAÇÃO:** {r["observacao"]}')
     st.warning('PENDENTE — será incluído no próximo Inventário Rotativo.')
  with te:
   if not encerrados:st.info('Nenhuma inconsistência encerrada.')
   for r in encerrados:
    with st.container(border=True):
     a,b,c=st.columns([1.2,2.2,1.5]);a.markdown(f'**{r["id"]}**');b.markdown(f'**{r["codigo"]}** · {r["descricao"]}');c.markdown(f'**{r["endereco"]}**')
     st.caption(f'Criado em {r["criado_em"]} · Encerrado em {r.get("encerrado_em") or "—"} · Equipe: {r["equipe"]}')
     st.write(f'**OBSERVAÇÃO:** {r["observacao"]}')
     st.success(f'TRATADO NO INVENTÁRIO: {r.get("inventario_doc") or "—"}')

# User administration
elif active=='Usuários' and st.session_state.profile=='Admin':
 st.subheader('Usuários');st.caption('Cadastro e controle de acesso dos usuários do sistema.')
 db_admin=firebase_db()
 if db_admin is None:
  st.error('Não foi possível acessar o Firestore. Verifique os Secrets do Firebase Admin.')
 else:
  with st.expander('NOVO USUÁRIO',True):
   with st.form('new_user_form'):
    c1,c2=st.columns(2)
    with c1: new_email=st.text_input('E-mail',placeholder='usuario@empresa.com')
    with c2: new_name=st.text_input('Nome',placeholder='Nome do usuário')
    c3,c4=st.columns(2)
    with c3: new_password=st.text_input('Senha inicial',type='password',placeholder='mínimo 6 caracteres')
    with c4: new_role=st.selectbox('Perfil',['OPERADOR','GESTOR','ADMIN'])
    create_user=st.form_submit_button('CRIAR USUÁRIO',type='primary',icon=':material/person_add:')
   if create_user:
    em=(new_email or '').strip().lower(); nm=(new_name or '').strip()
    try:
     if not em or not new_password or len(new_password)<6: st.error('Informe e-mail e senha com pelo menos 6 caracteres.')
     else:
      u=firebase_auth.create_user(email=em,password=new_password,display_name=nm or em.split('@')[0])
      db_admin.collection('usuarios').document(u.uid).set({'email':em,'nome':nm or em.split('@')[0],'perfil':new_role,'ativo':True,'criado_em':firestore.SERVER_TIMESTAMP})
      st.success('Usuário criado com sucesso.'); st.rerun()
    except Exception as e:
     msg=str(e)
     if 'EMAIL_EXISTS' in msg or 'already exists' in msg.lower(): msg='Este e-mail já está cadastrado.'
     st.error('Não foi possível criar o usuário: '+msg)
  st.markdown('### USUÁRIOS CADASTRADOS')
  try:
   users=list(firebase_auth.list_users().iterate_all())
  except Exception as e:
   st.error('Não foi possível listar os usuários: '+str(e)); users=[]
  for u in users:
   snap=db_admin.collection('usuarios').document(u.uid).get(); d=snap.to_dict() if snap.exists else {}
   with st.container(border=True):
    a,b,c,dcol=st.columns([2.3,1.4,1.1,1.0])
    with a: st.write('**'+(d.get('nome') or u.display_name or 'Sem nome')+'**'); st.caption(u.email or '')
    roles=['OPERADOR','GESTOR','ADMIN']; current=str(d.get('perfil','OPERADOR')).upper(); current=current if current in roles else 'OPERADOR'
    with b: role=st.selectbox('Perfil',roles,index=roles.index(current),key='role_'+u.uid)
    with c: active_user=st.checkbox('Ativo',value=not u.disabled,key='active_'+u.uid)
    with dcol:
     if st.button('SALVAR',key='save_'+u.uid,icon=':material/save:'):
      try:
       firebase_auth.update_user(u.uid,disabled=not active_user)
       db_admin.collection('usuarios').document(u.uid).set({'perfil':role,'ativo':active_user,'email':u.email or '','nome':d.get('nome') or u.display_name or '','atualizado_em':firestore.SERVER_TIMESTAMP},merge=True)
       st.success('Atualizado.'); st.rerun()
      except Exception as e: st.error('Erro ao atualizar: '+str(e))

# Settings
elif active=='Configurações':
 st.subheader('Configurações');st.caption('Agora as configurações são gravadas em armazenamento local do aplicativo. Cada grupo abaixo é independente.')
 with st.expander('01 · TEMA',True):
  t=st.radio('Modo',['Dark','Clean'],index=0 if cfg['theme']=='Dark' else 1,horizontal=True)
  if t!=cfg['theme']:cfg['theme']=t;persist_cfg();st.rerun()
 with st.expander('02 · TIPOGRAFIA',True):
  fonts=['Arial','Inter','Roboto','Poppins','Montserrat','Georgia','Verdana','Trebuchet MS'];a,b=st.columns(2);cfg['font']=a.selectbox('Tipo de letra',fonts,index=fonts.index(cfg['font']));cfg['font_size']=b.slider('Tamanho geral',12,22,cfg['font_size']);cfg['title_size']=st.slider('Tamanho do título',22,48,cfg['title_size'])
 with st.expander('03 · CORES',True):
  a,b=st.columns(2);cfg['primary']=a.color_picker('Cor principal',cfg['primary']);cfg['hover']=b.color_picker('Cor ao passar o mouse',cfg['hover']);keys=['dark_bg','dark_panel','dark_panel2','dark_border','dark_text','dark_muted'] if cfg['theme']=='Dark' else ['clean_bg','clean_panel','clean_panel2','clean_border','clean_text','clean_muted'];labs=['Fundo','Painéis','Painel secundário','Bordas','Texto','Texto secundário'];a,b=st.columns(2)
  for i,k in enumerate(keys):
   cfg[k]=(a if i%2==0 else b).color_picker(labs[i],cfg[k],key='cp_'+k)
  # widgets acima atualizam seus próprios valores no próximo rerun; abaixo usamos uma forma direta para os seis campos
  for k in keys:
   if k not in cfg:cfg[k]=DEFAULT[k]
 with st.expander('04 · LOGO',True):
  f=st.file_uploader('Logo da empresa',type=['png','jpg','jpeg','webp','svg'],key='logo_up')
  if f:st.session_state.logo=(f.getvalue(),f.name);save('logo',st.session_state.logo)
  a,b,c=st.columns(3);cfg['logo_w']=a.slider('Largura',80,320,cfg['logo_w']);cfg['logo_h']=b.slider('Altura',40,180,cfg['logo_h']);cfg['logo_align']=c.selectbox('Alinhamento',['left','center','right'],index=['left','center','right'].index(cfg['logo_align']));cfg['logo_top']=st.slider('Subir / descer logo',-100,100,cfg['logo_top'])
  if st.session_state.logo[0]:st.image(st.session_state.logo[0],width=min(cfg['logo_w'],320));
  if st.button('Remover logo'):st.session_state.logo=(None,'');save('logo',st.session_state.logo);st.rerun()
 with st.expander('05 · ELEMENTOS DA LATERAL — INDEPENDENTES',True):
  cfg['sidebar_sub']=st.text_input('Texto abaixo da logo',cfg['sidebar_sub']);cfg['sub_top']=st.slider('Posição do subtítulo',-60,100,cfg['sub_top']);cfg['menu']=st.text_input('Título do menu',cfg['menu']);cfg['menu_top']=st.slider('Posição do título MENU',-60,100,cfg['menu_top']);cfg['gap']=st.slider('Espaço antes do MENU',0,60,cfg['gap']);cfg['menu_gap']=st.slider('Espaçamento entre botões do menu',0,40,cfg.get('menu_gap',8));cfg['sidebar_align']=st.selectbox('Alinhamento dos tópicos',['left','center','right'],index=['left','center','right'].index(cfg['sidebar_align']));cfg['sidebar_font']=st.slider('Tamanho dos tópicos',10,22,cfg['sidebar_font']);cfg['item_h']=st.slider('Altura dos tópicos',30,70,cfg['item_h']);cfg['icon_color']=st.color_picker('Cor dos ícones',cfg['icon_color'])
  a,b=st.columns(2);cfg['dash_top']=a.slider('Dashboard — posição',-30,50,cfg['dash_top']);cfg['inv_top']=b.slider('Inventário — posição',-30,50,cfg['inv_top']);a,b=st.columns(2);cfg['db_top']=a.slider('Banco — posição',-30,50,cfg['db_top']);cfg['reg_top']=b.slider('Registro — posição',-30,50,cfg['reg_top']);a,b=st.columns(2);cfg['report_top']=a.slider('Reportar Inconsistências — posição',-30,50,cfg.get('report_top',0));cfg['settings_top']=b.slider('Configurações — posição',-30,50,cfg['settings_top'])
  st.caption('Cada tópico pode ter sua própria posição vertical.')
  for k,l in [('dash','Dashboard'),('inv','Inventário Rotativo'),('db','Banco de Dados'),('reg','Registro'),('report','Reportar Inconsistências'),('settings','Configurações')]:cfg[k]=st.text_input(l,cfg[k],key='menu_'+k)
 with st.expander('06 · TEXTOS DAS PÁGINAS',False):
  for k,l in [('title','Título principal'),('subtitle','Subtítulo principal'),('dashboard_title','Título Dashboard'),('inventory_title','Título Inventário'),('database_title','Título Banco'),('register_title','Título Registro')]:cfg[k]=st.text_input(l,cfg.get(k,DEFAULT.get(k,'')),key='pg_'+k)
 with st.expander('07 · INVENTÁRIO',True):cfg['blind_default']=st.checkbox('Contagem cega por padrão',cfg['blind_default'])
 if st.button('SALVAR TODAS AS CONFIGURAÇÕES',type='primary',use_container_width=True):persist_cfg();st.success('Configurações salvas.');st.rerun()
 if st.button('RESTAURAR PADRÃO'):st.session_state.cfg=DEFAULT.copy();persist_cfg();st.rerun()
